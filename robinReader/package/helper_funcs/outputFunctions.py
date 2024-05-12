import pandas as pd
import openpyxl as px
from openpyxl.worksheet.table import Table, TableStyleInfo
from io import BytesIO
from aws_jserver import upload_fileobj_to_s3
import os

def writeCSV(tradeList, object_name):
    df = pd.DataFrame(tradeList)
    
    # Rename headers
    df.columns.values[0:15] = [
        "Ticker",
        "Contract Description",
        "# Cons",
        "Avg Buy",
        "Total Buy",
        "Avg Sell",
        "Total Sell",
        "% Change",
        "Net Change",
        "Buy Date",
        "Buy DoW",
        "Sell Date",
        "Sell DoW",
        "Days Held",
        "Let Exp?"
    ]

    # Create a BytesIO object
    output = BytesIO()

    # Save the DataFrame to an Excel writer
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='welcome', index=False)
        writer.close()

    # Load the workbook from the BytesIO object
    output.seek(0)
    wb = px.load_workbook(output)
    ws = wb.active

    # Create a table and apply styles as before
    tab = Table(displayName="Table1", ref=ws.dimensions)
    style = TableStyleInfo(name="TableStyleMedium23", showRowStripes=True, showColumnStripes=False)
    tab.tableStyleInfo = style

    # Iterate over all columns and adjust their widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Style the '% Change' column with percent format
    number_format = '0.00%'
    for cell in ws['H']:
        cell.number_format = number_format

    # Style the buys/sells columns with currency($) formats
    number_format = '_($* #,##0.00_);_($* (#,##0.00);_($* "-"??_);_(@_)'
    for cell in ws['D']:
        cell.number_format = number_format
    
    for cell in ws['E']:
        cell.number_format = number_format

    for cell in ws['F']:
        cell.number_format = number_format

    for cell in ws['G']:
        cell.number_format = number_format

    # Style the net column with currency format that shows red for negative values
    number_format = '$#,##0.00_);[Red]($#,##0.00)'
    for cell in ws['I']:
        cell.number_format = number_format

    ws.add_table(tab)
    
    # Save workbook back to BytesIO object
    output = BytesIO()
    wb.save(output)
    wb.close()

    # Reset file pointer to the beginning after saving
    output.seek(0)

    # Upload the BytesIO object to S3
    if upload_fileobj_to_s3(output, os.environ['XLSX_BUCKET'], object_name):
        return True
    return False